# -*- coding: utf-8 -*-
"""
KPI Parc Conteneurs - Navis N4
==============================
Interroge l'API N4 (authentification HTTP Basic), calcule les statistiques
et genere un classeur Excel : donnees filtrables + tableaux + graphiques natifs.

Usage :
    python kpi_stats.py
    python kpi_stats.py --user MON_USER --url "http://SERVEUR:PORT/apex/api/query?filtername=units_1&operatorId=OPERATEUR&complexId=COMPLEXE&facilityId=TERMINAL&yardId=PARC"
    python kpi_stats.py --user X --password Y --out mon_rapport.xlsx
    python kpi_stats.py --in response.xml        (hors ligne, a partir d'un fichier deja telecharge)

Le mot de passe est demande de facon masquee s'il n'est pas fourni.
"""

import argparse
import getpass
import os
import sys
import io
import datetime
import xml.etree.ElementTree as ET
from collections import Counter, OrderedDict

DEFAULT_URL = ("http://SERVEUR:PORT/apex/api/query?filtername=units_1"
               "&operatorId=OPERATEUR&complexId=COMPLEXE&facilityId=TERMINAL&yardId=PARC")


def load_dotenv():
    """Charge un fichier .env situe a cote du script (sans dependance externe).
    Format : CLE=valeur, une par ligne, # pour les commentaires.
    Les variables deja definies dans l'environnement systeme ont la priorite."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, val = line.split("=", 1)
            key, val = key.strip(), val.strip().strip('"').strip("'")
            os.environ.setdefault(key, val)

# seuil d'aging (meme unite que la colonne Dwell)
AGING_THRESHOLD = 120


# --------------------------------------------------------------------------- I/O
def fetch_api(url, user, password):
    """Recupere le XML via HTTP Basic Auth. requests si dispo, sinon urllib."""
    try:
        import requests
        from requests.auth import HTTPBasicAuth
        r = requests.get(url, auth=HTTPBasicAuth(user, password), timeout=120)
        r.raise_for_status()
        return r.text
    except ImportError:
        import urllib.request
        import base64
        req = urllib.request.Request(url)
        tok = base64.b64encode(("%s:%s" % (user, password)).encode()).decode()
        req.add_header("Authorization", "Basic " + tok)
        with urllib.request.urlopen(req, timeout=120) as resp:
            return resp.read().decode("utf-8", "replace")


def parse_xml(text):
    """Retourne (colonnes, lignes) depuis le XML query-response."""
    cols, rows = [], []
    in_cols = False
    cur = None
    for ev, el in ET.iterparse(io.StringIO(text), events=("start", "end")):
        if ev == "start":
            if el.tag == "columns":
                in_cols = True
            elif el.tag == "row":
                cur = []
        else:
            if el.tag == "column" and in_cols:
                cols.append((el.text or "").strip())
            elif el.tag == "columns":
                in_cols = False
            elif el.tag == "field" and cur is not None:
                cur.append((el.text or "").strip())
            elif el.tag == "row":
                rows.append(cur)
                cur = None
                el.clear()
    return cols, rows


# ----------------------------------------------------------------- agregations
class Data:
    def __init__(self, cols, rows):
        self.cols = cols
        self.rows = rows
        self.idx = {c: i for i, c in enumerate(cols)}

    def has(self, name):
        return name in self.idx

    def get(self, row, name):
        j = self.idx.get(name, -1)
        return row[j] if 0 <= j < len(row) else ""

    def count_by(self, name, top=None):
        if not self.has(name):
            return []
        j = self.idx[name]
        c = Counter((r[j] if j < len(r) else "") or "(vide)" for r in self.rows)
        items = c.most_common()
        return items[:top] if top else items

    def count_true(self, name):
        if not self.has(name):
            return 0
        j = self.idx[name]
        return sum(1 for r in self.rows if j < len(r) and r[j].lower() == "true")

    def numeric(self, name):
        if not self.has(name):
            return []
        j = self.idx[name]
        out = []
        for r in self.rows:
            if j < len(r):
                try:
                    out.append(float(r[j]))
                except ValueError:
                    pass
        return out


def percentile(sorted_vals, p):
    if not sorted_vals:
        return 0
    k = int(len(sorted_vals) * p)
    return sorted_vals[min(k, len(sorted_vals) - 1)]


# --------------------------------------------------------------- console report
def print_summary(d):
    n = len(d.rows)
    line = "-" * 56
    print(line)
    print("  KPI PARC CONTENEURS  -  %d conteneurs" % n)
    print(line)

    def block(title, pairs):
        print("\n" + title)
        for k, v in pairs:
            bar = ""
            if isinstance(v, int) and n:
                bar = "  " + "#" * int(v * 30.0 / n)
            print("   %-32s %8s%s" % (k, ("{:,}".format(v) if isinstance(v, int) else v), bar))

    block("Categorie", d.count_by("Category"))
    block("Plein / Vide", d.count_by("Frght Kind"))
    block("Top armateurs", d.count_by("Line Op", 8))
    block("Etat transit", d.count_by("T-State"))

    print("\nSpecifiques")
    print("   %-32s %8s" % ("Reefers (Reqs Power)", "{:,}".format(d.count_true("Reqs Power"))))
    print("   %-32s %8s" % ("Dangereux (Hazardous?)", "{:,}".format(d.count_true("Hazardous?"))))
    print("   %-32s %8s" % ("Hors gabarit (Is OOG)", "{:,}".format(d.count_true("Is OOG"))))
    print("   %-32s %8s" % ("Bloques navire (Stop-Vsl)", "{:,}".format(d.count_true("Stop-Vsl"))))
    print("   %-32s %8s" % ("Bloques route (Stop-Road)", "{:,}".format(d.count_true("Stop-Road"))))

    dv = d.numeric("Dwell")
    if dv:
        dv.sort()
        aging = sum(1 for x in dv if x > AGING_THRESHOLD)
        print("\nTemps de sejour (Dwell)")
        print("   mediane %.0f | moyenne %.0f | p90 %.0f | max %.0f"
              % (dv[len(dv) // 2], sum(dv) / len(dv), percentile(dv, 0.90), dv[-1]))
        print("   aging > %d : %d conteneurs (%d%%)"
              % (AGING_THRESHOLD, aging, round(aging * 100.0 / n)))
    print(line)


# --------------------------------------------------------------- Excel report
def build_excel(d, out_path):
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    n = len(d.rows)
    NAVY = "12314F"
    HEADER = PatternFill("solid", fgColor=NAVY)
    HFONT = Font(color="FFFFFF", bold=True)
    THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    # ---------- Feuille Resume ----------
    ws = wb.active
    ws.title = "Resume"
    ws["A1"] = "KPI Parc Conteneurs - Navis N4"
    ws["A1"].font = Font(size=15, bold=True, color=NAVY)
    ws["A2"] = "Genere le %s  -  %d conteneurs" % (
        datetime.datetime.now().strftime("%d/%m/%Y %H:%M"), n)
    ws["A2"].font = Font(size=10, color="808080")

    cat = dict(d.count_by("Category"))
    fk = dict(d.count_by("Frght Kind"))
    dv = sorted(d.numeric("Dwell"))
    aging = sum(1 for x in dv if x > AGING_THRESHOLD)
    kpis = [
        ("Total conteneurs", n),
        ("Import", cat.get("Import", 0)),
        ("Export", cat.get("Export", 0)),
        ("En parc (Storage)", cat.get("Storage", 0)),
        ("Plein (FCL)", fk.get("FCL", 0) + fk.get("Full", 0)),
        ("Vide", fk.get("Empty", 0)),
        ("Reefers", d.count_true("Reqs Power")),
        ("Dangereux", d.count_true("Hazardous?")),
        ("Hors gabarit", d.count_true("Is OOG")),
        ("Bloques navire", d.count_true("Stop-Vsl")),
        ("Bloques route", d.count_true("Stop-Road")),
        ("Dwell median", int(dv[len(dv) // 2]) if dv else 0),
        ("Dwell p90", int(percentile(dv, 0.90)) if dv else 0),
        ("Dwell max", int(dv[-1]) if dv else 0),
        ("Aging > %d" % AGING_THRESHOLD, aging),
    ]
    r0 = 4
    ws.cell(r0, 1, "Indicateur").fill = HEADER
    ws.cell(r0, 1).font = HFONT
    ws.cell(r0, 2, "Valeur").fill = HEADER
    ws.cell(r0, 2).font = HFONT
    ws.cell(r0, 3, "% du parc").fill = HEADER
    ws.cell(r0, 3).font = HFONT
    for i, (k, v) in enumerate(kpis):
        ws.cell(r0 + 1 + i, 1, k)
        ws.cell(r0 + 1 + i, 2, v)
        if "Dwell" not in k and n:
            ws.cell(r0 + 1 + i, 3, round(v * 100.0 / n)).number_format = '0"%"'
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

    # ---------- Feuille Analyses (tableaux + graphiques) ----------
    wa = wb.create_sheet("Analyses")

    def write_table(title, pairs, anchor_row, anchor_col=1):
        wa.cell(anchor_row, anchor_col, title).font = Font(bold=True, color=NAVY)
        wa.cell(anchor_row + 1, anchor_col, "Valeur").fill = HEADER
        wa.cell(anchor_row + 1, anchor_col).font = HFONT
        wa.cell(anchor_row + 1, anchor_col + 1, "Nombre").fill = HEADER
        wa.cell(anchor_row + 1, anchor_col + 1).font = HFONT
        for i, (k, v) in enumerate(pairs):
            wa.cell(anchor_row + 2 + i, anchor_col, k)
            wa.cell(anchor_row + 2 + i, anchor_col + 1, v)
        return anchor_row + 1, anchor_row + 1 + len(pairs)  # header_row, last_row

    def bar_chart(title, hr, lr, at):
        ch = BarChart()
        ch.type = "bar"
        ch.title = title
        ch.height = 7
        ch.width = 13
        ch.legend = None
        data = Reference(wa, min_col=2, min_row=hr, max_row=lr)
        cats = Reference(wa, min_col=1, min_row=hr + 1, max_row=lr)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        wa.add_chart(ch, at)

    def pie_chart(title, hr, lr, at):
        ch = PieChart()
        ch.title = title
        ch.height = 7
        ch.width = 8
        data = Reference(wa, min_col=2, min_row=hr, max_row=lr)
        cats = Reference(wa, min_col=1, min_row=hr + 1, max_row=lr)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        wa.add_chart(ch, at)

    # dwell buckets
    edges = [0, 6, 12, 24, 48, 72, 120, 168, 240, 360, 10 ** 9]
    labels = ["<6", "6-12", "12-24", "24-48", "48-72", "72-120",
              "120-168", "168-240", "240-360", ">360"]
    buckets = [0] * (len(edges) - 1)
    for x in dv:
        for i in range(len(edges) - 1):
            if edges[i] <= x < edges[i + 1]:
                buckets[i] += 1
                break
    dwell_pairs = list(zip(labels, buckets))

    hr1, lr1 = write_table("Par categorie", d.count_by("Category"), 1, 1)
    bar_chart("Par categorie", hr1, lr1, "D2")
    hr2, lr2 = write_table("Plein / Vide", d.count_by("Frght Kind"), 1, 12)
    pie_chart("Plein / Vide", hr2, lr2, "O2")
    hr3, lr3 = write_table("Top 12 armateurs", d.count_by("Line Op", 12), 16, 1)
    bar_chart("Top 12 armateurs", hr3, lr3, "D16")
    hr4, lr4 = write_table("Top 12 POD", d.count_by("POD", 12), 16, 12)
    bar_chart("Top 12 POD", hr4, lr4, "O16")
    hr5, lr5 = write_table("Etat transit", d.count_by("T-State"), 34, 1)
    bar_chart("Etat transit", hr5, lr5, "D34")
    hr6, lr6 = write_table("Dwell - tranches", dwell_pairs, 34, 12)
    bar_chart("Dwell - tranches", hr6, lr6, "O34")
    for col in "ABMN":
        wa.column_dimensions[col].width = 20

    # ---------- Feuille Donnees (filtrable) ----------
    wd = wb.create_sheet("Donnees")
    for j, c in enumerate(d.cols, 1):
        cell = wd.cell(1, j, c)
        cell.fill = HEADER
        cell.font = HFONT
    for i, row in enumerate(d.rows, 2):
        for j, v in enumerate(row, 1):
            # convertit les nombres pour permettre tri/filtre numerique
            wd.cell(i, j, _num(v))
    wd.freeze_panes = "A2"
    wd.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(d.cols)), len(d.rows) + 1)
    for j in range(1, len(d.cols) + 1):
        wd.column_dimensions[get_column_letter(j)].width = 15

    wb.save(out_path)


def _num(v):
    if v is None or v == "":
        return ""
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except ValueError:
        return v


# ------------------------------------------------------------------------ main
def main():
    load_dotenv()  # lit .env (N4_USER, N4_PASSWORD, N4_URL) s'il existe

    ap = argparse.ArgumentParser(description="KPI Parc Conteneurs - Navis N4")
    ap.add_argument("--url", default=os.environ.get("N4_URL", DEFAULT_URL), help="URL de l'API N4")
    ap.add_argument("--user", help="identifiant (Basic Auth)")
    ap.add_argument("--password", help="mot de passe (sinon demande masquee)")
    ap.add_argument("--out", help="fichier Excel de sortie")
    ap.add_argument("--in", dest="infile", help="lire un fichier XML au lieu de l'API")
    args = ap.parse_args()

    if args.infile:
        print("Lecture du fichier %s ..." % args.infile)
        with open(args.infile, encoding="utf-8", errors="replace") as f:
            text = f.read()
    else:
        # priorite : argument CLI  >  .env / variable d'environnement  >  saisie clavier
        user = args.user or os.environ.get("N4_USER") or input("Identifiant N4 : ").strip()
        password = (args.password or os.environ.get("N4_PASSWORD")
                    or getpass.getpass("Mot de passe : "))
        if os.environ.get("N4_USER"):
            print("Identifiant charge depuis .env : %s" % user)
        print("Appel de l'API ...")
        try:
            text = fetch_api(args.url, user, password)
        except Exception as e:
            print("ERREUR appel API : %s" % e, file=sys.stderr)
            print("Verifie l'URL, le reseau du terminal, et tes identifiants.", file=sys.stderr)
            sys.exit(1)

    cols, rows = parse_xml(text)
    if not rows:
        print("Aucune donnee recue. Reponse inattendue ?", file=sys.stderr)
        sys.exit(1)
    d = Data(cols, rows)

    print_summary(d)

    out = args.out or ("kpi_stats_%s.xlsx" % datetime.datetime.now().strftime("%Y%m%d_%H%M"))
    print("\nGeneration du classeur Excel ...")
    build_excel(d, out)
    print("OK -> %s  (%d lignes, feuilles : Resume / Analyses / Donnees)" % (out, len(rows)))


if __name__ == "__main__":
    main()
